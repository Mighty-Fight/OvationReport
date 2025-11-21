import os
import re
import shutil
import win32com.client
from collections import Counter
from openpyxl import load_workbook
from dotenv import load_dotenv

load_dotenv()

def convertir_docx_a_pdf_y_extraer_numeros(carpeta_entrada, carpeta_salida, archivo_txt_salida, carpeta_errores, carpeta_repetidos):
    word = win32com.client.Dispatch("Word.Application")
    word.Visible = False

    os.makedirs(carpeta_salida, exist_ok=True)
    os.makedirs(carpeta_errores, exist_ok=True)
    os.makedirs(carpeta_repetidos, exist_ok=True)

    numeros_ot = set()
    archivos_no_validos = []
    archivos_repetidos = []
    total_archivos_docx = 0
    total_convertidos = 0

    # --- Primer barrido: contar OTs ---
    lista_ot = []
    for archivo in os.listdir(carpeta_entrada):
        if archivo.endswith(".docx") and not archivo.startswith("~$"):
            match = re.search(r'OT_(\d+)_INFORME', archivo)
            if match:
                lista_ot.append(match.group(1))

    conteo_ot = Counter(lista_ot)
    ot_repetidos = {ot for ot, count in conteo_ot.items() if count > 1}

    # --- Segundo barrido: procesar archivos ---
    for archivo in os.listdir(carpeta_entrada):
        if archivo.endswith(".docx") and not archivo.startswith("~$"):
            total_archivos_docx += 1
            origen = os.path.join(carpeta_entrada, archivo)

            match = re.search(r'OT_(\d+)_INFORME', archivo)
            
            # === CASO 1: Formato inválido ===
            if not match:
                destino_docx = os.path.join(carpeta_errores, archivo)
                destino_pdf = os.path.join(
                    carpeta_errores, os.path.splitext(archivo)[0] + ".pdf"
                )
                shutil.move(origen, destino_docx)
                print(f"❌ Formato inválido, movido: {archivo}")
                try:
                    doc = word.Documents.Open(destino_docx)
                    doc.SaveAs(destino_pdf, FileFormat=17)
                    doc.Close()
                    archivos_no_validos.append(archivo)
                    print(f"📄 Convertido a PDF (mal formato): {destino_pdf}")
                    total_convertidos += 1
                except Exception as e:
                    print(f"⚠️ Error al convertir archivo con mal formato {archivo}: {e}")
                continue

            ot_num = match.group(1)

            # === CASO 2: OT repetido ===
            if ot_num in ot_repetidos:
                destino_docx = os.path.join(carpeta_repetidos, archivo)
                destino_pdf = os.path.join(
                    carpeta_repetidos, os.path.splitext(archivo)[0] + ".pdf"
                )
                shutil.move(origen, destino_docx)
                print(f"♻️ Movido a repetidos: {archivo}")
                try:
                    doc = word.Documents.Open(destino_docx)
                    doc.SaveAs(destino_pdf, FileFormat=17)
                    doc.Close()
                    archivos_repetidos.append(archivo)
                    print(f"📄 Convertido a PDF (repetido): {destino_pdf}")
                    total_convertidos += 1
                    numeros_ot.add(ot_num)
                except Exception as e:
                    print(f"⚠️ Error al convertir archivo repetido {archivo}: {e}")
                continue

            # === CASO 3: Válido y único ===
            destino_pdf = os.path.join(carpeta_salida, os.path.splitext(archivo)[0] + ".pdf")
            print(f"Convirtiendo: {archivo} → {destino_pdf}")
            try:
                doc = word.Documents.Open(origen)
                doc.SaveAs(destino_pdf, FileFormat=17)
                doc.Close()
                total_convertidos += 1
                numeros_ot.add(ot_num)
            except Exception as e:
                print(f"❌ Error al convertir {archivo}: {e}")

    word.Quit()

    # --- Procesar Excel SOLO HOJA1 ---
    numeros_ordenados = sorted(numeros_ot.union(ot_repetidos))
    ruta_excel = r"C:\Users\CamiloAmaya\Documents\Prueba_excel\Libro1.xlsx"

    try:
        wb = load_workbook(ruta_excel)
        hoja1 = wb["Hoja1"]

        # Limpiar completamente Hoja1
        for fila in range(1, hoja1.max_row + 1):
            for col in range(1, 9):
                hoja1.cell(row=fila, column=col, value=None)

        # Escribir encabezado
        hoja1["A1"] = "Numero OT"

        # Escribir solo la lista de OTs
        for i, ot in enumerate(numeros_ordenados, start=2):
            hoja1[f"A{i}"] = int(ot)    
            hoja1[f"A{i}"].number_format = '0'  # Formato número sin decimales

        wb.save(ruta_excel)
        print(f"✅ Excel actualizado SOLO con números de OT en Hoja1: {ruta_excel}")

    except Exception as e:
        print(f"❌ Error al escribir en Excel: {e}")

    # --- Crear archivo TXT ---
    cadena_formateada = ' OR '.join(f'"{num}"' for num in numeros_ordenados)
    with open(archivo_txt_salida, 'w', encoding='utf-8') as f:
        f.write(cadena_formateada)

    # --- Resultados ---
    print("\n✅ Números de OT exportados a archivo:")
    print(archivo_txt_salida)

    if archivos_no_validos:
        print("\n⚠️ Archivos con nombre inválido convertidos y movidos:")
        for archivo in archivos_no_validos:
            print(f" - {archivo}")

    if archivos_repetidos:
        print("\n⚠️ Archivos repetidos convertidos y movidos:")
        for archivo in archivos_repetidos:
            print(f" - {archivo}")

    print("\n📊 RESUMEN FINAL:")
    print(f" - Total de archivos .docx encontrados: {total_archivos_docx}")
    print(f" - Archivos convertidos exitosamente: {total_convertidos}")
    print(f" - Archivos con nombre inválido: {len(archivos_no_validos)}")
    print(f" - Archivos repetidos: {len(archivos_repetidos)}")


# === RUTAS DESDE ENV ===
carpeta_entrada = os.getenv("carpeta_entrada")
carpeta_salida = os.getenv("carpeta_salida")
archivo_txt_salida = os.getenv("archivo_txt_salida")
carpeta_errores = os.getenv("carpeta_errores")
carpeta_repetidos = os.getenv("carpeta_repetidos")
ruta_excel = os.getenv("ruta_excel")

print("📂 Entrada:", carpeta_entrada)
print("📂 Salida:", carpeta_salida)
print("📂 TXT:", archivo_txt_salida)
print("📂 Errores:", carpeta_errores)
print("📂 Repetidos:", carpeta_repetidos)
print("📂 Excel:", ruta_excel)

convertir_docx_a_pdf_y_extraer_numeros(
    carpeta_entrada,
    carpeta_salida,
    archivo_txt_salida,
    carpeta_errores,
    carpeta_repetidos
)

print("\nProceso completado ✅")
