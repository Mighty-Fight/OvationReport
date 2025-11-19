import pandas as pd
from openpyxl import load_workbook

# Rutas de los archivos
file1 = r"C:\Users\CamiloAmaya\Downloads\PAUTOMATECAPAResults326.xlsx"
file2 = r"C:\Users\CamiloAmaya\Documents\update_fechas\NEXXO_UPDATE_FECHAS.xlsx"

# Cargar archivo origen (resultados)
df1 = pd.read_excel(file1, dtype=str)  # todo como texto para evitar errores
df1 = df1.fillna("")  # reemplazar NaN por vacío

# Cargar archivo destino en la hoja JULIO
book = load_workbook(file2)
sheet = book["JULIO"]

# Encontrar índice de la tabla (Tabla1)
# --- Asumimos que Tabla1 empieza en la fila 1 (ajustar si empieza más abajo)
df2 = pd.read_excel(file2, sheet_name="JULIO", dtype=str)
df2 = df2.fillna("")

# Crear un diccionario desde file1: Columna A -> Columna H
mapa = dict(zip(df1.iloc[:, 0], df1.iloc[:, 7]))  # A=0, H=7

# Recorrer filas de la hoja destino y actualizar Columna I
for row in range(2, sheet.max_row + 1):  # desde la fila 2 (salta encabezado)
    valor_colA = str(sheet.cell(row=row, column=1).value).strip()
    if valor_colA in mapa:
        sheet.cell(row=row, column=9).value = mapa[valor_colA]  # Col I = 9

# Guardar cambios
book.save(file2)

print("Actualización completada ✅")